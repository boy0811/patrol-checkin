from flask import Blueprint, request, send_file, render_template, redirect, session, flash, jsonify
from io import StringIO, BytesIO
from datetime import datetime, timedelta, timezone

import csv

from models import db, Record, Member, Point, Report

admin_records_bp = Blueprint('admin_records', __name__, url_prefix='/admin')

# ✅ 匯出表單頁面
@admin_records_bp.route('/export_form')
def export_form():
    if not session.get('admin'):
        return redirect('/login')
    return render_template('admin/export_form.html')

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ✅ 匯出 XLSX 功能
@admin_records_bp.route('/export_records', methods=['POST'])
def export_records():
    if not session.get('admin'):
        return redirect('/login')

    start_date_str = request.form.get('start_date')
    end_date_str = request.form.get('end_date')

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        end_date = datetime.combine(end_date, datetime.max.time())
    except ValueError:
        flash("⚠️ 日期格式錯誤", "danger")
        return redirect('/admin/export_form')

    records = Record.query.filter(
        Record.timestamp >= start_date,
        Record.timestamp <= end_date
    ).all()

    # ✅ 建立 Excel 檔
    wb = Workbook()
    ws = wb.active
    ws.title = "簽到紀錄"

    # ✅ 表頭
    headers = ['隊員姓名', '級職', '巡邏點', '簽到時間']
    ws.append(headers)

    # ✅ 資料列
    for r in records:
        ws.append([
            r.member.name,
            r.member.title,
            r.point.name,
            r.timestamp.strftime('%Y-%m-%d %H:%M:%S')
        ])

    # ✅ 自動調整欄寬
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # ✅ 儲存到 BytesIO
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'records_{start_date_str}_to_{end_date_str}.xlsx'
    )


# ✅ 簽到紀錄頁面
@admin_records_bp.route('/records')
def view_records():
    if not session.get('admin'):
        return redirect('/login')
    return render_template('admin/records.html')

# ✅ 提供 JS API 用於載入簽到紀錄
@admin_records_bp.route('/api/records')
def api_records():
    if not session.get('admin'):
        return jsonify([])

    records = Record.query.order_by(Record.timestamp.desc()).all()
    result = []
    taipei_time = timezone(timedelta(hours=8))  # 台灣時區

    for r in records:
        timestamp = r.timestamp
        if timestamp.tzinfo is None:
            timestamp = timestamp.replace(tzinfo=timezone.utc)
        local_time = timestamp.astimezone(taipei_time).strftime('%Y-%m-%d %H:%M:%S')

        result.append({
            'account': r.member.account,
            'name': r.member.name,
            'point': r.point.name,
            'timestamp': local_time
        })
    return jsonify(result)

# ✅ 顯示緊急通報紀錄頁面
@admin_records_bp.route('/reports')
def view_reports():
    if not session.get('admin'):
        return redirect('/login')

    reports = db.session.query(
        Member.account,             # ✅ 修正為 account
        Member.name,
        Report.description,
        Report.photo_filename,
        Report.timestamp
    ).join(Member, Report.member_id == Member.id).order_by(Report.timestamp.desc()).all()

    return render_template('admin/reports.html', reports=reports)

# ✅ 匯出緊急通報 XLSX
@admin_records_bp.route('/export_reports', methods=['POST'])
def export_reports():
    if not session.get('admin'):
        return redirect('/login')

    start_date_str = request.form.get('start_date')
    end_date_str = request.form.get('end_date')

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        end_date = datetime.combine(end_date, datetime.max.time())
    except ValueError:
        flash("⚠️ 日期格式錯誤", "danger")
        return redirect('/admin/export_form')

    reports = db.session.query(
        Member.account,
        Member.name,
        Report.description,
        Report.photo_filename,
        Report.timestamp
    ).join(Member, Report.member_id == Member.id)\
     .filter(Report.timestamp >= start_date, Report.timestamp <= end_date)\
     .order_by(Report.timestamp.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "緊急通報"

    # ✅ 表頭
    headers = ['帳號', '姓名', '通報內容', '照片檔名', '通報時間']
    ws.append(headers)

    # ✅ 資料列
    for report in reports:
        account, name, desc, photo, timestamp = report
        ws.append([
            account,
            name,
            desc,
            photo,
            timestamp.strftime('%Y-%m-%d %H:%M:%S')
        ])

    # ✅ 自動調整欄寬
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'reports_{start_date_str}_to_{end_date_str}.xlsx'
    )
